import logging

from openpyxl.styles import PatternFill


class Task1:
    def __init__(self, excel):
        self.excel = excel
        self.name_column = 0
        self.header_row_number = 2
        self.find_name_column(self.excel.hr_sheet)
        self.find_reference_month(self.excel.hr_sheet)
        if not self.excel.hr_month_cell:
            logging.error("Nem találtam meg a referencia hónap oszlopát: '" + self.excel.hr_month_column + "'")
            return
        logging.info('Referencia hónap oszlopát megtaláltam: ' + self.excel.hr_month_cell.column_letter)
        self.process_people()
        self.attempt_to_color_promoted_cells()

    def process_people(self):
        logging.info('Indítom az emberek feldolgozását...')
        for row in self.excel.full_sheet.iter_rows():
            if 'neve' in row[0].value:
                continue
            name = row[0].value
            self.find_and_color_name_in_hr(self.excel.hr_sheet, name.splitlines()[0])

    def attempt_to_color_promoted_cells(self):
        logging.info(self.excel.SMALL_LINE + self.excel.SMALL_LINE)
        logging.info('Feldolgozás befejezve.')
        logging.info('A következő cellákat készítettem elő színezésre: ')
        for person in self.excel.cells_promoted_for_coloring:
            logging.info(self.excel.SMALL_ARROW + ' ' + person + ': ' + self.excel.cells_promoted_for_coloring[person].coordinate)
        self.color_promoted_cells()

    def color_promoted_cells(self):
        color = "FF99FF"
        logging.info(self.excel.SMALL_LINE + self.excel.SMALL_LINE)
        logging.info('Színezés engedélyezve. A szín: ' + color)
        pinkFill = PatternFill(start_color=color, fill_type="solid")
        for person in self.excel.cells_promoted_for_coloring:
            logging.info("Színezem: " + self.excel.cells_promoted_for_coloring[person].coordinate)
            self.excel.cells_promoted_for_coloring[person].fill = pinkFill
        logging.info('Színezés sikeresen befejzve!')

    def find_reference_month(self, hr_sheet):
        self.excel.hr_month_column = str(self.excel.year) + '_' + str(self.excel.month_number)
        logging.info("Referencia hónap oszlop értéke: '" + self.excel.hr_month_column + "'")
        logging.info('Keresem a referencia hónap oszlopot...')
        for cell in hr_sheet[self.header_row_number]:
            if cell.value == self.excel.hr_month_column:
                self.excel.hr_month_cell = cell

    def find_and_color_name_in_hr(self, hr_sheet, name):
        logging.info(self.excel.SMALL_LINE)
        logging.info('Célszemély: ' + name)
        found_coordinates = []
        target_cell_found = None
        for row in hr_sheet.iter_rows():
            if row[self.name_column].value == name:
                found_coordinate = row[self.name_column].coordinate
                logging.info('Megtaláltam egy sort: ' + found_coordinate)
                found_coordinates.append(found_coordinate)
                reference_month_cell = row[self.excel.hr_month_cell.column - 1]
                if isinstance(reference_month_cell.value, int) and reference_month_cell.value == 1:
                    target_cell_found = reference_month_cell
                    break
        if target_cell_found is None:
            msg = "Valami nem stimmel, nem találtam megfelelő sort '" + name + "'-hoz/-hez. A vizsgált sorok: " + str(
                found_coordinates)
            logging.error(msg)
            raise Exception(msg)
        logging.info("Megtaláltam a megfelelő cellát '" + name + "'-hoz/-hez: " + target_cell_found.coordinate)
        self.excel.cells_promoted_for_coloring[name] = target_cell_found

    def find_name_column(self, hr_sheet):
        for cell in hr_sheet[self.header_row_number]:
            if cell.value == 'Név':
                self.name_column = cell.column - 1
