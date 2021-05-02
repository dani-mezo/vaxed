import logging
import random

from openpyxl import load_workbook

from message import fun, LONG_LINE
from sheet_data import SheetData
from task import Task


class ExcelProcessor:
    def __init__(self, app, terv, full, year, month, month_number, mean_duration, target_color, tasks):
        logging.info('Megkezdjük a feldolgozást. Ez eltarthat egy darabig, türelem! ... ' + random.choice(fun))
        logging.info('Átlagosan ennyi másodpercig tart: ' + str(format(float(mean_duration), '.2f')))
        Task.i = 0
        self.terv_workbook = load_workbook(filename=terv)
        self.full_workbook = load_workbook(filename=full)
        self.full = full
        self.terv = terv
        self.app = app
        self.year = year
        self.month = month
        self.month_number = month_number
        self.target_color = target_color
        self.tasks = tasks
        self.cells_promoted_for_coloring = dict()
        self.month_number = str(self.month_number) if int(self.month_number) >= 10 else '0' + str(self.month_number)
        self.header_month_column_name = str(self.year) + '_' + str(self.month_number)
        self.error_cells = []
        cut_year = year[2:]
        full_workbook_referred_sheet_names = list(filter(lambda sheet_name: cut_year in sheet_name and month in sheet_name,
                                                    self.full_workbook.get_sheet_names()))
        logging.debug('Betöltöttem a FULL táblából a következő füleket: ' + str(full_workbook_referred_sheet_names))
        if len(full_workbook_referred_sheet_names) > 1:
            logging.error('Valami nem stimmel, több fül is megfelelt a kiválasztásnak: ' + str(full_workbook_referred_sheet_names))
            return
        if len(full_workbook_referred_sheet_names) < 1:
            logging.error("Valami nem stimmel, nem találok fület ami megfelel az év '" + year + "', hónap '" + month + "' feltételeknek.")
            return
        try:
            full_sheet_name = full_workbook_referred_sheet_names[0]
            full_sheet = self.full_workbook.get_sheet_by_name(full_sheet_name)
            full_sheet_data = SheetData(self.full_workbook, full_sheet, 'FULL - ' + full_sheet_name, False)
            self.do_tasks(full_sheet_data, full_sheet)
            self.close()
        except Exception as e:
            logging.error("Probléma merült fel, a végrehajtást megszakítom.")
            logging.exception(e)
            self.close()
            return

    def do_tasks(self, full_sheet_data, full_sheet):
        for i, task in enumerate(self.tasks, start=1):
            if task is None:
                continue
            full_task_name = next(iter(self.tasks[i - 1].keys()))
            terv_task_name = next(iter(self.tasks[i - 1].values()))
            just_verify_presence = True if 'HR' in terv_task_name else False
            task = Task(self.terv_workbook, full_sheet_data, full_sheet, full_task_name, terv_task_name,
                        self.header_month_column_name, just_verify_presence, self.target_color, i)
            self.error_cells.extend(task.failed_verification_cells)
            self.save()

    def save(self):
        logging.info('Mentés: ' + self.full)
        self.full_workbook.save(self.full)
        logging.info('Mentés: ' + self.terv)
        self.terv_workbook.save(self.terv)
        logging.info(LONG_LINE)

    def close(self):
        if self.terv_workbook is not None:
            self.terv_workbook.close()
        if self.full_workbook is not None:
            self.full_workbook.close()




